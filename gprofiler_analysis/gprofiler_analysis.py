
## Developed by Carl McIntosh
## Date: 5 March 2025
## Version 0.5.0

## Python Imports
from gprofiler import GProfiler
from openpyxl import load_workbook
import argparse
import csv
import json
import os
import pandas as pd
import sys
import subprocess
import glob

## Make Gene List Directories
def make_dir():
	gene_list_dir = 'GeneLists'
	os.makedirs(gene_list_dir, exist_ok=True)

	return gene_list_dir

def print_output_files():
	print("\n############################### Output File List ###############################")
	list_output_files = glob.glob("./GeneLists/*GeneList*")
	subprocess.run(["ls","-l"] + sorted(list_output_files))



## Generate individual text files for data
def print_excel_sheet_to_csv(EXCEL_FILE_PATH):
	print("\n####################### Exporting Sheets from Excel File #######################")

	sheets_list = []

	#### Now extract Excel Sheets to
	wb = load_workbook(EXCEL_FILE_PATH)

	for idx, excel_sheet_name in enumerate(wb.sheetnames):
		print(str(idx + 1) +  '. ' + excel_sheet_name + ".csv")
		excel_sheet_wb = wb[excel_sheet_name]
		sheets_list.append(excel_sheet_name + ".csv")

		with open(excel_sheet_name + ".csv", 'w', newline='', encoding='utf-8') as MY_FILE:
			csv_writer = csv.writer(MY_FILE)
			for row in excel_sheet_wb.rows:
				csv_writer.writerow([cell.value for cell in row])
	return sheets_list

## Select Sheet
def select_sheet_by_number(sheets_list):
	print("\n------------------------------- Available Sheets -------------------------------")
	for idx, excel_sheet_name in enumerate(sheets_list):
		print(str(idx + 1) + '. ' + sheets_list[idx])

	sheet_number = int(input("\nSelect Sheet by Number: "))
	print("You selected: " + str(sheet_number) + ". '" +  sheets_list[sheet_number - 1] + "'\n")
	return (sheet_number - 1)

def select_group_column_by_number(df):
	column_names = df.columns.tolist()
	for idx, excel_sheet_name in enumerate(column_names):
		print(str(idx + 1) + '. ' + column_names[idx])

	column_number = int(input("Select Group Column by Number: "))
	print("You selected: " + str(column_number) + ". '" +  column_names[column_number - 1] + "'\n")
	return column_names[column_number - 1]

def select_genes_column_by_number(df):
	column_names = df.columns.tolist()
	for idx, excel_sheet_name in enumerate(column_names):
		print(str(idx + 1) + '. ' + column_names[idx])

	column_number = int(input("Select Gene Names Column by Number: "))
	print("You selected: " + str(column_number) + ". '" +  column_names[column_number - 1] + "'\n")
	return column_names[column_number - 1]


def select_column_by_number(df, column_selector, column_id):
	if column_id not in df.columns.tolist():
		print("\n######## Select " + column_selector + " Column Name ########")
		column_names = df.columns.tolist()
		for idx, excel_sheet_name in enumerate(column_names):
			print(str(idx + 1) + '. ' + column_names[idx])

		column_number = int(input("Select " + column_selector + " Column by Number: "))
		print("You selected: " + str(column_number) + ". '" +  column_names[column_number - 1] + "'\n")

		column_id = column_names[column_number - 1]

	return column_id

def stats_from_unfiltered(df, group_number, group_column_name, adj_pvalue_header):
	tmp_df = df[df[group_column_name] == group_number]
	filename_list = None
	try:
		filename_list = tmp_df["Exon_annotation"].sort_values().unique().tolist()
	except:
		filename_list = tmp_df["filename"].sort_values().unique().tolist()

	filename = filename_list[0].replace(".csv","")

	max_adj_pvalue = tmp_df[adj_pvalue_header].max()
	min_adj_pvalue = tmp_df[adj_pvalue_header].min()

	return (filename, min_adj_pvalue, max_adj_pvalue)

def run_gprofiler_analysis_R(in_args):

	gprofiler_analysis_R_script_name = "gprofiler_analysis.R"
	gprofiler_analysis_R_script = in_args.script_dir + "/" + gprofiler_analysis_R_script_name
	out_dir = os.path.abspath(in_args.out)

	if os.path.exists(gprofiler_analysis_R_script):
		print(f"File '{gprofiler_analysis_R_script_name}' exists.")

		if in_args.use_bkg_known:
			subprocess.run(["Rscript", gprofiler_analysis_R_script, "--output", "./", "--use_bkg_known", "TRUE"])
		else:
			subprocess.run(["Rscript", gprofiler_analysis_R_script, "--output", "./"])
	else:
		print(f"File '{gprofiler_analysis_R_script_name}' does not exist.")

######## get_args - start ##########################################################################
def get_args():
	"""*get_args* - parses program's arg values.

	Returns
	arguments (*dict*) - user selected arguments parser.parse_args()

	"""
	parser = argparse.ArgumentParser(
		prog='gprofiler_analysis.py',
		description="Description: This script 'gprofiler_analysis.py' a wrapper app for g:Profiler.\nSee included documentation file 'gprofiler_analysis.md' for details")

	#### Parameters
	parser.add_argument("-v","--version", action='version', version='%(prog)s version: Version 1.0.0 - Feb 2025')
	parser.add_argument("-o", "--out", help="Location of output directory.\nIf not specified, files will be written to the current working directory.", default='./', required=False)
	parser.add_argument("--adj_pval_filter", help="Adjusted p-value filter (default=0.05)). ", type=float, default=0.05, required=False)

	group_gprofiler = parser.add_argument_group("gProfiler arguments")
	parser.add_argument("-r","--run_gprofiler", help="Run gProfiler.", action='store_true', default=False)
	parser.add_argument("--use_bkg_known", help="Also include Known Background.", action='store_true', default=False)

	group_required = parser.add_argument_group("required arguments")
	group_required.add_argument("-f", "--excel_file", help="Path to Excel file created by scanpy.get.rank_genes_groups_df(). Or, a txt file for tab-separated values or a csv file for comma-separated values.", required=True)

	group_required.add_argument("--lfc_cutoff", help="Absolute value for LFC cut off value (float). ", type=float, required=True)



	return parser.parse_args()

######## get_args - end ############################################################################

def get_link(in_description, in_live_link):

	link = str('\t<li><a href="' + in_live_link + '" target="_blank">' + in_description + '</a></li>\n')
	return link

def get_relative_link(in_relative_path):

	description = in_relative_path.replace("./results_custom_bkg/","")
	description = description.replace("_bkg_index.html","")

	relative_link = str('\t<li><a href="' + in_relative_path + '" target="_blank">' + description + '</a></li>\n')
	return relative_link

def summarize_rprofiler(in_args, stats_table_df):
	# in_args.results_custom_bkg

	#### Links to Reports

	gprofiler_html_output_files = glob.glob("./results_custom_bkg/*.html")

	custom_bg_results_count = len(gprofiler_html_output_files)

	combine_terms_df = None
	combine_terms_column_count = 0

	url_dictionary = {}

	with open ("custom_urls.txt", "r") as URLS_FILE:
		URL_LINES = URLS_FILE.readlines()
		for index, line in enumerate(URL_LINES):
			if index > 0:
				line = line.replace('"','')
				data = line.split(",")
				url_dictionary[data[1]] = data[2]


	with open ("gProfiler_summary_results.html", "w") as SUMMARY_HTML:

		SUMMARY_HTML.write('<!DOCTYPE html>\n<html lang="en">\n<head>\n\t<meta charset="utf-8" />\n\t<title>Summary gProfiler Results</title>\n</head>\n<body>\n')

		lfc_cutoff_down_regulated = -1 * in_args.lfc_cutoff
		lfc_cutoff_up_regulated   = in_args.lfc_cutoff

		# Summarize Parameters Selected.
		SUMMARY_HTML.write("<h2>Parameters Selected</h2>\n")
		SUMMARY_HTML.write("<p>Location of Excel File: " + in_args.excel_file + "</p>\n")
		SUMMARY_HTML.write("<p>Location of Output Directory: " + in_args.cwd + "</p>\n")
		SUMMARY_HTML.write("<p>Adjusted pValue Cutoff: Cutoff < " + str(in_args.adj_pval_filter) + "</p>\n")
		SUMMARY_HTML.write("<p>Select Down Regulated genes with LFC Cut Off: LFC < " + str(lfc_cutoff_down_regulated) + "</p>\n")
		SUMMARY_HTML.write("<p>Select Up Regulated genes with LFC Cut Off: LFC > " + str(lfc_cutoff_up_regulated) + "</p>\n")
		SUMMARY_HTML.write("<p>Filtered Unique Gene IDs (Custom Background Set): " + str(in_args.filtered_uniq_genes_count) + "</p>\n")

		#### Gene Lists Summary
		SUMMARY_HTML.write("<h2>Gene Lists Summary</h2>\n")
		stats_table_string = stats_table_df.to_html()
		SUMMARY_HTML.write("\n" + stats_table_string + "\n")

		## Live Links
		SUMMARY_HTML.write("<h2>Live Links to gProfiler</h2>\n")
		SUMMARY_HTML.write("<ol>\n")
		for index, key in enumerate(url_dictionary.keys()):
			SUMMARY_HTML.write(get_link(key, url_dictionary[key]))

		SUMMARY_HTML.write("</ol>\n")


		SUMMARY_HTML.write("<h2>Static Links to gProfiler</h2>\n")
		SUMMARY_HTML.write("<ol>\n")


		# for index, html_relative_path in enumerate(sorted(gprofiler_html_output_files)):
		for index, key in enumerate(url_dictionary.keys()):
			key=key.replace(":","_")
			html_relative_path="./results_custom_bkg/" + key.replace("-","_") + "_bkg_custom.html"
			SUMMARY_HTML.write(get_relative_link(html_relative_path))

			if False:
				csv_file = html_relative_path.replace(".html",".csv")
				csv_file = csv_file.replace("./","")

				print("Adding: " + csv_file)
				if index == 0:
					print(csv_file)
					combine_terms_df = pd.read_csv(csv_file, sep=',', header=1)
					combine_terms_column_count = combine_terms_df.shape[1]

				else:
					new_df = pd.read_csv(csv_file, sep=',', header=1)
					if (combine_terms_column_count == new_df.shape[1]):
						combine_terms_df =pd.concat([combine_terms_df,new_df], axis=0)
					else:
						print ("Column count didn't match: " + csv_file)

		print(combine_terms_df)
		SUMMARY_HTML.write("</ol>\n")

		## Close
		SUMMARY_HTML.write("\n</body>\n</html>\n")



######## main - start ##############################################################################
def main():
	"""*main* - main function.

	"""
	######## Parse arguments ########
	args = get_args()

	args.script_dir = os.path.dirname(sys.argv[0])
	args.script_dir = os.path.abspath(args.script_dir)

	# args to values
	pvals_adj_filter=args.adj_pval_filter

	## Test files
	print("################################################################################")
	EXCEL_FILE_PATH=args.excel_file

	#### Set and Print Current Working Directory
	WORKING_DIR=os.path.abspath(args.out)
	os.chdir(WORKING_DIR)
	args.cwd = os.getcwd()

	gene_list_dir = make_dir()

	Group_All_File_CSV = open(gene_list_dir + "/Group_All_GeneList.csv", "w")

	lfc_cutoff_down_regulated = -1 * args.lfc_cutoff
	lfc_cutoff_up_regulated   = args.lfc_cutoff

	# Summarize Parameters Selected.
	print("\n-------- Parameters Selected ------")
	print("Location of Excel File:\n\t" + EXCEL_FILE_PATH)
	print("\nLocation of Output Directory:\n\t" + args.cwd)
	print("\nAdjusted pValue Cutoff:\n\tCutoff < " + str(pvals_adj_filter))
	print("\nSelect Down Regulated genes with LFC Cut Off:\n\tLFC < " + str(lfc_cutoff_down_regulated))
	print("\nSelect Up Regulated genes with LFC Cut Off:\n\tLFC > " + str(lfc_cutoff_up_regulated))

	if args.run_gprofiler:
		print("\n-------- gProfiler Selected -------")
		print("gProfiler script will be run (https://biit.cs.ut.ee/gprofiler/gost).")
		if args.use_bkg_known:
			print("\nTwo gProfiler results will be generated:")
			print("    1. custom background in directory 'results_custom_bkg'")
			print("    2. all known genes background in directory 'results_know_bkg'\n")
		else:
			print("\ngProfiler will run against custom set only. ")
			print("    1. custom background in directory 'results_custom_bkg'\n")
		print("For information see:\n\thttps://biit.cs.ut.ee/gprofiler/page/docs#statistical_domain_scope")
	else: print("\ngProfiler script will not be run. Use --run_gprofiler to run.\n")

	## Keep Count
	gene_list_count = {}

	gene_list_count["pvals_adj_filter"]=pvals_adj_filter

	# #### Input Excel Speadsheet
	# EXCEL_FILE=os.path.basename(EXCEL_FILE_PATH)
	# print("Location of Excel Spreadsheet: \n\t" + EXCEL_FILE_PATH)

	# Placeholder
	unfiltered_df = None

	#### Query Dictionary
	query_dictionary={}

	if EXCEL_FILE_PATH.endswith(".xlsx"):
		## Export Sheets to csv file
		sheets_list = print_excel_sheet_to_csv(EXCEL_FILE_PATH)

		## Select Sheet to explore
		sheet_number = select_sheet_by_number(sheets_list)

		# Read Data and Clean Dataframe
		print("############################## Pre-Filter Summary ##############################")
		unfiltered_df = pd.read_csv(sheets_list[sheet_number], header=0)
		print(unfiltered_df)
	elif EXCEL_FILE_PATH.endswith(".txt"):
		unfiltered_df = pd.read_csv(EXCEL_FILE_PATH, sep='\t', header=0)
		print(unfiltered_df)
	elif EXCEL_FILE_PATH.endswith(".csv"):
		unfiltered_df = pd.read_csv(EXCEL_FILE_PATH, sep=',', header=0)
		print(unfiltered_df)
	else:
		print("\n!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
		print("ERROR: Input file must with extention xlsx, txt or csv")
		print("This program will terminate.\n")
		sys.exit()

	# unfiltered_df = unfiltered_df.drop(unfiltered_df.columns[[0]], axis=1)
	gene_list_count['total_row_count'] = len(unfiltered_df)
	print("Total Row Count for All Groups: " + str(gene_list_count['total_row_count']))


	# Select Gene ID column
	gene_names_col_name = select_column_by_number(unfiltered_df, "Gene ID", "Gene")

	## Select Group Column by Number
	group_column_name = select_column_by_number(unfiltered_df, "Group ID", "group")

	## Select logfold_col Column by Number
	logfold_col = select_column_by_number(unfiltered_df, "Log Fold", "log2FoldChange")

	## Select logfold_col Column by Number
	padj_column_name = select_column_by_number(unfiltered_df, "Adj. p-Value", "padj")

	all_uniq_genes = unfiltered_df[gene_names_col_name].sort_values().unique().tolist()
	all_uniq_genes_count = len(all_uniq_genes)
	args.all_uniq_genes_count = all_uniq_genes_count
	print("Unfiltered Unique Gene IDs: " + str(all_uniq_genes_count))

	with open(gene_list_dir + "/bkg_all_GeneList.txt", "w") as outfile:
		outfile.write("\n".join(all_uniq_genes) + "\n")

	## Filter dataframe
	print("\n####################### Filter Summary for Adj. p-Value ######################")
	pvals_adj_filter_df = unfiltered_df[unfiltered_df[padj_column_name] < pvals_adj_filter]
	gene_list_count['pvals_adj_gene_count'] = len(pvals_adj_filter_df)
	print("Filtered Row Count (pvals_adj < " + str(pvals_adj_filter) +") for all groups: " + str(gene_list_count['pvals_adj_gene_count']))

	max_filtered_pvals_adj = pvals_adj_filter_df[padj_column_name].max()
	min_filtered_pvals_adj = pvals_adj_filter_df[padj_column_name].min()
	print("Range of filter pvals_adj: [" + str(min_filtered_pvals_adj) + ", " + str(max_filtered_pvals_adj) + "]")

	## Get list of all genes
	filtered_uniq_genes = pvals_adj_filter_df[gene_names_col_name].sort_values().unique().tolist()
	filtered_uniq_genes_count = len(filtered_uniq_genes)
	args.filtered_uniq_genes_count = filtered_uniq_genes_count
	print("Filtered Unique Gene IDs: " + str(filtered_uniq_genes_count))
	with open(gene_list_dir + "/bkg_filtered_GeneList.txt", "w") as outfile:
		outfile.write("\n".join(filtered_uniq_genes) + "\n")
	# Group_All_File_CSV.write("background;" + ",".join(filtered_uniq_genes) + "\n")
	Group_All_File_CSV.write("background;" + ",".join(filtered_uniq_genes) + "\n")

	query_dictionary={}
	# query_dictionary['organism'] = 'hsapiens'
	# query_dictionary['sources']=['GO:MF','GO:CC','GO:BP','KEGG','REAC','WP','TF','MIRNA','HPA','CORUM','HP']
	query_dictionary['user_threshold']=1e-8
	# query_dictionary['ordered']=True
	query_dictionary['highlight']=True
	query_dictionary['query']={}

	all_groups_query={}
	gene_list_count["group_count"] = pvals_adj_filter_df[group_column_name].max()

	stats_table = {
		"Group Id":[],
		"Filename":[],
		"Filtered Genes":[],
		"Depleted Genes":[],
		"Enriched Genes":[],
		"Min Unfiltered padj":[],
		"Max Unfiltered padj":[],
		"Gene List File":[]
	}

	## Create All Genelist file for for gProfiler
	Group_All_File = open(gene_list_dir + "/Group_All_GeneList.txt", "w")

	for group_number in range(0, gene_list_count["group_count"] + 1):
		group_df = pvals_adj_filter_df[pvals_adj_filter_df[group_column_name] == group_number]

		## Get some values from unfiltered
		(filename, min_adj_pvalue, max_adj_pvalue) = stats_from_unfiltered(unfiltered_df, group_number, group_column_name, padj_column_name)

		# Thresholds for logfold:
		# < -1.5 : depleted
		# . > 1.5 : enriched


		lfc_cutoff_up_regulated

		group_Depleted = group_df[group_df[logfold_col]  < lfc_cutoff_down_regulated]
		group_enriched =  group_df[group_df[logfold_col] >  lfc_cutoff_up_regulated]

		group_name = "Group_" + str(group_number) + "_" + filename

		gene_list_count[group_name + " Total Gene X Group"]          = len(group_df)
		gene_list_count[group_name + " Depleted Gene Count"]          = len(group_Depleted)
		gene_list_count[group_name + " Enriched Gene Count"]           =  len(group_enriched)

		stats_table["Group Id"].append(group_number)
		stats_table["Filename"].append(filename)
		stats_table["Filtered Genes"].append(gene_list_count[group_name + " Total Gene X Group"])
		stats_table["Depleted Genes"].append(gene_list_count[group_name + " Depleted Gene Count"])
		stats_table["Enriched Genes"].append(gene_list_count[group_name + " Enriched Gene Count"])
		stats_table["Min Unfiltered padj"].append(min_adj_pvalue)
		stats_table["Max Unfiltered padj"].append(max_adj_pvalue)

		# all_groups[group_name] = group_df[gene_names_col_name].to_list()

		outfile = None

		if False:
			outfile = open(gene_list_dir + '/Group_' + str(group_number) + "_GeneList.txt", "w")

		###
		combo_out_file_name = gene_list_dir + '/Group_' + str(group_number) + "_GeneListCombo.txt"
		stats_table["Gene List File"].append(combo_out_file_name.replace("GeneLists/",""))
		combo_out = open(combo_out_file_name, "w")
		combo_out.write("> " + group_name + "\n")



		query_dictionary['query']["depleted"] = []
		query_dictionary['query']["enriched"] = []

		if gene_list_count[group_name + " Depleted Gene Count"] > 0:
			query_dictionary['query']["depleted"] = group_Depleted[gene_names_col_name].to_list()
			if False:
				outfile.write("> "+ 'Group_' + str(group_number) + " Depleted" + "\n")
				outfile.write("\n".join(query_dictionary['query']["depleted"]) + "\n")
				Group_All_File.write("> "+ 'Group_' + str(group_number) + " Depleted" + "\n")
				Group_All_File.write("\n".join(query_dictionary['query']["depleted"]) + "\n")

			combo_out.write("\n".join(query_dictionary['query']["depleted"]) + "\n")

		if gene_list_count[group_name + " Enriched Gene Count"] > 0:
			query_dictionary['query']["enriched"] = group_enriched[gene_names_col_name].to_list()
			if False:
				outfile.write("> "+ 'Group_' + str(group_number) + " Enriched" + "\n")
				outfile.write("\n".join(query_dictionary['query']["enriched"]) + "\n")
				Group_All_File.write("> "+ 'Group_' + str(group_number) + " Enriched" + "\n")
				Group_All_File.write("\n".join(query_dictionary['query']["enriched"]) + "\n")

			combo_out.write("\n".join(query_dictionary['query']["enriched"]) + "\n")

			# Group_All_File_CSV.write(group_name + "_Depleted:" +  ",".join(query_dictionary['query']["depleted"]) + "\n")
			# Group_All_File_CSV.write(group_name + "_Enriched:" +  ",".join(query_dictionary['query']["enriched"]) + "\n")

		if False:
			outfile.close()

		if gene_list_count[group_name + " Depleted Gene Count"] > 0 or gene_list_count[group_name + " Enriched Gene Count"] > 0:
			Group_All_File_CSV.write(group_name + ";" + ",".join(query_dictionary['query']["depleted"] + query_dictionary['query']["enriched"]) + "\n")

		# with open(group_name + ".json", "w") as outfile:
		# 	json.dump(query_dictionary, outfile, indent=4)

		if False:
			# https://pypi.org/project/gprofiler-official/
			gp = GProfiler(return_dataframe=True)

			results = gp.profile(
				organism = query_dictionary['organism'],
				query = group_df[gene_names_col_name].to_list(),
				sources=query_dictionary['sources'],
				combined=True,
				user_threshold=query_dictionary['user_threshold']
			)

			results.to_csv(group_name + "_GeneListResults.csv", index=False)

	Group_All_File.close()
	Group_All_File_CSV.close()

	print("\n############################## Summary Information #############################")
	stats_table_df = pd.DataFrame(stats_table)
	print(stats_table_df.to_string())

	stats_table_df.to_csv('gene_counts_summary_table.txt', sep='\t', index=False)

	# print_output_files()

	if args.run_gprofiler: run_gprofiler_analysis_R(args)

	summarize_rprofiler(args, stats_table_df)


if __name__ == "__main__":
		main()
	######## main - end ################################################################################
